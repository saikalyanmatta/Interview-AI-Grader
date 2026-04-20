import io
from datetime import datetime
from fastapi import APIRouter, Request, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession
from db.database import get_db
from db.models import ScheduledInterview, InterviewCandidate, Interview, InterviewReport, Job, User
from routes.auth import get_current_user

router = APIRouter()


@router.get("/scheduled-interviews")
async def list_scheduled(request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    items = db.query(ScheduledInterview).filter(ScheduledInterview.employer_id == user["id"]).all()
    result = []
    for si in items:
        candidates = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si.id).all()
        attempts = db.query(Interview).filter(Interview.scheduled_interview_id == si.id).all()
        job_title = None
        if si.job_id:
            j = db.query(Job).filter(Job.id == si.job_id).first()
            job_title = j.title if j else None
        d = _si_to_dict(si)
        d["jobTitle"] = job_title
        d["candidateCount"] = len(candidates)
        d["attemptCount"] = len(attempts)
        result.append(d)
    return result


@router.post("/scheduled-interviews", status_code=201)
async def create_scheduled(request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    body = await request.json()
    title = body.get("title")
    start_time = body.get("startTime")
    deadline_time = body.get("deadlineTime")
    if not title or not start_time or not deadline_time:
        raise HTTPException(status_code=400, detail="title, startTime, deadlineTime are required")

    si = ScheduledInterview(
        employer_id=user["id"],
        job_id=body.get("jobId"),
        title=title,
        start_time=datetime.fromisoformat(start_time.replace("Z", "+00:00")),
        deadline_time=datetime.fromisoformat(deadline_time.replace("Z", "+00:00")),
        coding_questions_count=body.get("codingQuestionsCount", 0),
        role=body.get("role", "Software Engineer"),
        difficulty=body.get("difficulty", "Medium"),
        interview_style=body.get("interviewStyle", "Professional"),
        interview_type=body.get("interviewType", "Mixed"),
        coding_language=body.get("codingLanguage", "Candidate's Choice"),
        question_complexity=body.get("questionComplexity", "Moderate"),
    )
    db.add(si)
    db.commit()
    db.refresh(si)
    return _si_to_dict(si)


@router.get("/scheduled-interviews/{si_id}")
async def get_scheduled(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")
    candidates = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()
    job_title = None
    if si.job_id:
        j = db.query(Job).filter(Job.id == si.job_id).first()
        job_title = j.title if j else None
    result = _si_to_dict(si)
    result["jobTitle"] = job_title
    result["candidates"] = [_candidate_to_dict(c) for c in candidates]
    return result


@router.delete("/scheduled-interviews/{si_id}")
async def delete_scheduled(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(si)
    db.commit()
    return {"success": True}


@router.post("/scheduled-interviews/{si_id}/candidates")
async def add_candidates(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")
    body = await request.json()
    emails = body.get("emails", [])
    if not isinstance(emails, list):
        raise HTTPException(status_code=400, detail="emails array required")

    existing = {c.email.lower() for c in db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()}
    import re
    to_add = list({e.lower().strip() for e in emails if e and re.match(r"\S+@\S+\.\S+", e)} - existing)

    for email in to_add:
        db.add(InterviewCandidate(scheduled_interview_id=si_id, email=email))
    db.commit()

    all_candidates = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()
    return {"added": len(to_add), "candidates": [_candidate_to_dict(c) for c in all_candidates]}


@router.delete("/scheduled-interviews/{si_id}/candidates")
async def remove_candidate(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")
    body = await request.json()
    email = body.get("email", "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="email required")
    c = db.query(InterviewCandidate).filter(
        InterviewCandidate.scheduled_interview_id == si_id,
        InterviewCandidate.email == email
    ).first()
    if c:
        db.delete(c)
        db.commit()
    remaining = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()
    return {"candidates": [_candidate_to_dict(c) for c in remaining]}


@router.get("/scheduled-interviews/{si_id}/results")
async def get_results(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")
    if datetime.utcnow() < si.start_time:
        raise HTTPException(status_code=403, detail="Results not available until interview start time")

    candidates = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()
    attempts = db.query(Interview).filter(Interview.scheduled_interview_id == si_id).all()

    results = []
    for iv in attempts:
        report = db.query(InterviewReport).filter(InterviewReport.interview_id == iv.id).first()
        db_user = db.query(User).filter(User.id == iv.user_id).first()
        results.append({
            "interviewId": iv.id,
            "candidateName": iv.candidate_name or (db_user.first_name if db_user else "Unknown"),
            "email": db_user.email if db_user else "",
            "status": iv.status,
            "report": _report_to_dict(report) if report else None,
            "overallScore": report.overall_score if report else None,
            "recommendation": report.recommendation if report else None,
            "skillScores": report.skill_scores if report else [],
            "englishScore": report.english_score if report else None,
            "behavioralScore": report.behavioral_score if report else None,
            "confidenceScore": report.confidence_score if report else None,
            "codingScore": getattr(report, "coding_score", None) if report else None,
            "technicalScore": getattr(report, "technical_score", None) if report else None,
        })

    attempted_emails = {r["email"].lower() for r in results}
    not_attempted = [
        {"email": c.email, "candidateName": None, "status": "not_attempted", "report": None}
        for c in candidates if c.email.lower() not in attempted_emails
    ]

    return {"scheduledInterview": _si_to_dict(si), "attempted": results, "notAttempted": not_attempted}


@router.get("/scheduled-interviews/{si_id}/results/export")
async def export_results(si_id: int, request: Request, db: DBSession = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == si_id, ScheduledInterview.employer_id == user["id"]).first()
    if not si:
        raise HTTPException(status_code=404, detail="Not found")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    candidates = db.query(InterviewCandidate).filter(InterviewCandidate.scheduled_interview_id == si_id).all()
    attempts = db.query(Interview).filter(Interview.scheduled_interview_id == si_id).all()

    results = []
    for iv in attempts:
        report = db.query(InterviewReport).filter(InterviewReport.interview_id == iv.id).first()
        db_user = db.query(User).filter(User.id == iv.user_id).first()
        skill_scores = report.skill_scores if report else []
        results.append({
            "candidateName": iv.candidate_name or (db_user.first_name if db_user else "Unknown"),
            "email": db_user.email if db_user else "",
            "overallScore": report.overall_score if report else None,
            "englishScore": report.english_score if report else None,
            "behavioralScore": report.behavioral_score if report else None,
            "confidenceScore": report.confidence_score if report else None,
            "codingScore": getattr(report, "coding_score", None) if report else None,
            "technicalScore": getattr(report, "technical_score", None) if report else None,
            "skillScores": skill_scores,
            "recommendation": report.recommendation if report else None,
            "feedback": report.feedback if report else "",
        })

    all_skill_names = list({s["skill"] for r in results for s in r["skillScores"]})
    base_headers = ["Name", "Email", "Overall Score", "Behavioral Score", "Coding Score", "Technical Theory Score", "English Score", "Confidence Score"]
    skill_headers = [f"Skill: {s}" for s in all_skill_names]
    headers = base_headers + skill_headers + ["Recommendation", "Feedback"]

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    attempted_emails = {r["email"].lower() for r in results}
    not_attempted = [{"candidateName": "", "email": c.email} for c in candidates if c.email.lower() not in attempted_emails]

    strongly_hire = [r for r in results if r["recommendation"] == "hire" and (r["overallScore"] or 0) >= 80]
    hire = [r for r in results if r["recommendation"] == "hire" and (r["overallScore"] or 0) < 80]
    maybe = [r for r in results if r["recommendation"] == "maybe"]
    no_hire = [r for r in results if r["recommendation"] == "no_hire"]

    def add_sheet(name, rows, use_full_headers=True):
        ws = wb.create_sheet(title=name)
        hdrs = headers if use_full_headers else ["Name", "Email"]
        ws.append(hdrs)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 18

    def make_row(r):
        base = [r["candidateName"], r["email"], r["overallScore"], r["behavioralScore"], r["codingScore"], r["technicalScore"], r["englishScore"], r["confidenceScore"]]
        skills = [next((s["score"] for s in r["skillScores"] if s["skill"] == sn), "") for sn in all_skill_names]
        return base + skills + [r["recommendation"], r["feedback"]]

    add_sheet("Strongly Hire", [make_row(r) for r in strongly_hire])
    add_sheet("Hire", [make_row(r) for r in hire + maybe])
    add_sheet("Do Not Hire", [make_row(r) for r in no_hire])
    add_sheet("Not Attempted", [[r["candidateName"], r["email"]] for r in not_attempted], use_full_headers=False)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=results-{si_id}.xlsx"}
    )


@router.post("/scheduled-interviews/validate-access")
async def validate_access(request: Request, db: DBSession = Depends(get_db)):
    body = await request.json()
    si_id = body.get("scheduledInterviewId")
    email = body.get("email", "").lower().strip()
    if not si_id or not email:
        raise HTTPException(status_code=400, detail="scheduledInterviewId and email required")

    si = db.query(ScheduledInterview).filter(ScheduledInterview.id == int(si_id)).first()
    if not si:
        return {"allowed": False, "reason": "Interview not found"}

    now = datetime.utcnow()
    if now < si.start_time:
        return {"allowed": False, "reason": "Interview has not started yet", "startTime": si.start_time.isoformat()}
    if now > si.deadline_time:
        return {"allowed": False, "reason": "Interview deadline has passed"}

    candidate = db.query(InterviewCandidate).filter(
        InterviewCandidate.scheduled_interview_id == int(si_id),
        InterviewCandidate.email == email
    ).first()

    if not candidate:
        return {"allowed": False, "reason": "You are not registered for this interview"}

    return {"allowed": True, "scheduledInterview": _si_to_dict(si)}


def _si_to_dict(si: ScheduledInterview) -> dict:
    return {
        "id": si.id,
        "employerId": si.employer_id,
        "jobId": si.job_id,
        "title": si.title,
        "startTime": si.start_time.isoformat() if si.start_time else None,
        "deadlineTime": si.deadline_time.isoformat() if si.deadline_time else None,
        "codingQuestionsCount": si.coding_questions_count,
        "role": si.role,
        "difficulty": si.difficulty,
        "interviewStyle": si.interview_style,
        "interviewType": getattr(si, "interview_type", "Mixed"),
        "codingLanguage": getattr(si, "coding_language", "Candidate's Choice"),
        "questionComplexity": getattr(si, "question_complexity", "Moderate"),
        "createdAt": si.created_at.isoformat() if si.created_at else None,
    }


def _candidate_to_dict(c: InterviewCandidate) -> dict:
    return {
        "id": c.id,
        "scheduledInterviewId": c.scheduled_interview_id,
        "email": c.email,
        "addedAt": c.added_at.isoformat() if c.added_at else None,
    }


def _report_to_dict(r: InterviewReport) -> dict:
    return {
        "id": r.id,
        "interviewId": r.interview_id,
        "englishScore": r.english_score,
        "englishFeedback": r.english_feedback,
        "overallScore": r.overall_score,
        "confidenceScore": r.confidence_score,
        "confidenceNotes": r.confidence_notes,
        "behavioralScore": r.behavioral_score,
        "behavioralAnalysis": r.behavioral_analysis,
        "communicationAnalysis": r.communication_analysis,
        "answerQualityBreakdown": r.answer_quality_breakdown,
        "stutterAnalysis": r.stutter_analysis,
        "skillScores": r.skill_scores,
        "recommendation": r.recommendation,
        "feedback": r.feedback,
        "createdAt": r.created_at.isoformat() if r.created_at else None,
    }
